"""
Excel export utility for backtest results with IST datetime formatting.

Exports comprehensive trade details, equity curve, and metrics to Excel with proper formatting.
"""

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
from zoneinfo import ZoneInfo

logger = logging.getLogger(__name__)

# IST timezone
IST = ZoneInfo("Asia/Kolkata")


def convert_to_ist(df: pd.DataFrame, timestamp_cols: list) -> pd.DataFrame:
    """
    Convert UTC timestamp columns to IST.
    
    Args:
        df: DataFrame with timestamp columns
        timestamp_cols: List of column names containing timestamps
        
    Returns:
        DataFrame with timestamps converted to IST
    """
    df = df.copy()
    
    for col in timestamp_cols:
        if col in df.columns:
            # Convert to datetime if not already
            if not pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = pd.to_datetime(df[col], utc=True, errors='coerce')
            
            # Convert to IST
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                # Ensure UTC timezone
                if df[col].dt.tz is None:
                    df[col] = df[col].dt.tz_localize('UTC')
                
                # Convert to IST
                df[col] = df[col].dt.tz_convert(IST)
                
                # Remove timezone info for Excel compatibility
                df[col] = df[col].dt.tz_localize(None)
    
    return df


def format_excel_sheet(writer, sheet_name: str, df: pd.DataFrame, freeze_panes: tuple = (1, 0)):
    """
    Format Excel sheet with proper styling using openpyxl.
    
    Args:
        writer: ExcelWriter object
        sheet_name: Name of the sheet
        df: DataFrame to format
        freeze_panes: Tuple of (row, col) to freeze panes
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    worksheet = writer.sheets[sheet_name]
    
    # Header styling
    header_fill = PatternFill(start_color="D7E4BD", end_color="D7E4BD", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply header formatting
    for col_num, value in enumerate(df.columns.values, start=1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Set column widths
    for i, col in enumerate(df.columns, start=1):
        # Calculate max width
        max_len = max(
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
            len(str(col))
        ) + 2
        column_letter = worksheet.cell(row=1, column=i).column_letter
        worksheet.column_dimensions[column_letter].width = min(max_len, 50)
    
    # Freeze panes
    if freeze_panes:
        # openpyxl uses 1-indexed, freeze_panes is (row, col)
        freeze_cell = worksheet.cell(row=freeze_panes[0] + 1, column=freeze_panes[1] + 1)
        worksheet.freeze_panes = freeze_cell



def export_to_excel(
    run_dir: Path,
    output_path: Optional[Path] = None,
    include_positions: bool = True,
    include_selection: bool = False,
    use_reports_folder: bool = True
) -> Path:
    """
    Export backtest results to Excel with IST datetime formatting.
    
    Args:
        run_dir: Path to run directory containing CSV files
        output_path: Optional custom output path
        include_positions: Include positions sheet
        include_selection: Include contract selection sheet
        use_reports_folder: Save to reports/ folder instead of run directory (default: True)
        
    Returns:
        Path to created Excel file
    """
    if output_path is None:
        if use_reports_folder:
            # Save to reports folder with run_id as filename
            reports_dir = Path("reports")
            reports_dir.mkdir(exist_ok=True)
            run_id = run_dir.name
            output_path = reports_dir / f"{run_id}.xlsx"
        else:
            # Save to run directory
            output_path = run_dir / "results.xlsx"
    
    logger.info(f"Exporting backtest results to Excel: {output_path}")
    
    # Load data files
    trades_path = run_dir / "trades.csv"
    equity_path = run_dir / "equity_curve.csv"
    positions_path = run_dir / "positions.csv"
    selection_path = run_dir / "selection.csv"
    metrics_path = run_dir / "metrics.json"
    manifest_path = run_dir / "manifest.json"
    
    # Create Excel writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # 1. Summary Sheet
        summary_data = []
        
        # Load manifest
        if manifest_path.exists():
            import json
            with open(manifest_path, 'r') as f:
                manifest = json.load(f)
            
            summary_data.append(['Run ID', manifest.get('run_id', 'N/A')])
            summary_data.append(['Strategy', manifest.get('strategy', 'N/A')])
            summary_data.append(['Start Date', manifest.get('start', 'N/A')])
            summary_data.append(['End Date', manifest.get('end', 'N/A')])
            summary_data.append(['Bar Size', manifest.get('bar_size', 'N/A')])
            summary_data.append(['', ''])  # Blank row
        
        # Load metrics
        if metrics_path.exists():
            import json
            with open(metrics_path, 'r') as f:
                metrics = json.load(f)
            
            summary_data.append(['=== Performance Metrics ===', ''])
            summary_data.append(['Total Trades', metrics.get('total_trades', 0)])
            summary_data.append(['Total Return (%)', f"{metrics.get('total_return_pct', 0):.2f}"])
            summary_data.append(['Max Drawdown (%)', f"{metrics.get('max_drawdown_pct', 0):.2f}"])
            summary_data.append(['Sharpe Ratio', f"{metrics.get('sharpe', 0):.2f}"])
            summary_data.append(['Win Rate (%)', f"{metrics.get('win_rate_pct', 0):.2f}"])
            summary_data.append(['Avg Trade P&L (INR)', f"{metrics.get('avg_trade_pnl', 0):,.2f}"])
            summary_data.append(['Final Equity (INR)', f"{metrics.get('final_equity', 100000):,.2f}"])
        
        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        format_excel_sheet(writer, 'Summary', summary_df)
        
        # 2. Trades Sheet
        if trades_path.exists():
            trades_df = pd.read_csv(trades_path)
            
            if len(trades_df) > 0:
                # Convert timestamps to IST
                timestamp_cols = ['timestamp', 'entry_time', 'exit_time']
                trades_df = convert_to_ist(trades_df, timestamp_cols)
                
                # Create formatted trades DataFrame with required columns
                formatted_trades = pd.DataFrame()
                
                # Order ID (sequential)
                formatted_trades['Order ID'] = range(1, len(trades_df) + 1)
                
                # Entry Date and Time
                if 'entry_time' in trades_df.columns:
                    formatted_trades['Entry Date'] = trades_df['entry_time'].dt.date
                    formatted_trades['Entry Time (IST)'] = trades_df['entry_time'].dt.strftime('%H:%M:%S')
                
                # Exit Date and Time
                if 'exit_time' in trades_df.columns:
                    formatted_trades['Exit Date'] = trades_df['exit_time'].dt.date
                    formatted_trades['Exit Time (IST)'] = trades_df['exit_time'].dt.strftime('%H:%M:%S')
                
                # Option Symbol
                formatted_trades['Option Symbol'] = trades_df.get('symbol', '')
                
                # Option Type (CE/PE)
                if 'symbol' in trades_df.columns:
                    formatted_trades['Option Type'] = trades_df['symbol'].apply(
                        lambda x: 'CE' if 'CE' in str(x) else ('PE' if 'PE' in str(x) else '')
                    )
                
                # Strike Price
                if 'symbol' in trades_df.columns:
                    # Extract strike from symbol (e.g., NIFTY02SEP2524650CE -> 24650)
                    formatted_trades['Strike'] = trades_df['symbol'].apply(
                        lambda x: int(''.join(filter(str.isdigit, str(x).split('CE')[0].split('PE')[0][-5:]))) 
                        if pd.notna(x) else 0
                    )
                
                # Predicted Direction (LONG for option buying)
                formatted_trades['Predicted Direction'] = 'LONG'
                
                # Entry Price
                formatted_trades['Entry Price'] = trades_df.get('entry_price', 0).round(2)
                
                # Exit Price
                formatted_trades['Exit Price'] = trades_df.get('exit_price', 0).round(2)
                
                # PnL (Points) - difference in premium
                if 'entry_price' in trades_df.columns and 'exit_price' in trades_df.columns:
                    formatted_trades['Pnl (Points)'] = (trades_df['exit_price'] - trades_df['entry_price']).round(2)
                
                # Gross P&L (before costs)
                if 'qty' in trades_df.columns:
                    qty = trades_df.get('qty', 1)
                    lot_size = 75  # NIFTY lot size
                    formatted_trades['Gross P&L'] = (
                        (trades_df.get('exit_price', 0) - trades_df.get('entry_price', 0)) * qty * lot_size
                    ).round(2)
                
                # Transaction Cost (commission + fees)
                # Assuming 20 per contract for entry + exit
                formatted_trades['Transaction Cost'] = 40.0
                
                # Net P&L (after costs)
                formatted_trades['Net P&L'] = trades_df.get('realized_pnl', 0).round(2)
                
                # Exit Reason - infer from PnL
                if 'realized_pnl' in trades_df.columns:
                    def infer_exit_reason(pnl, pnl_points):
                        """Infer exit reason from PnL"""
                        if abs(pnl_points) < 0.5:  # Very small change
                            return 'FLAT'
                        elif pnl > 0:  # Profit
                            return 'TARGET'
                        else:  # Loss
                            return 'STOP_LOSS'
                    
                    # Calculate PnL points for inference
                    pnl_points = (trades_df.get('exit_price', 0) - trades_df.get('entry_price', 0))
                    formatted_trades['Exit Reason'] = [
                        infer_exit_reason(pnl, points) 
                        for pnl, points in zip(trades_df.get('realized_pnl', 0), pnl_points)
                    ]
                else:
                    formatted_trades['Exit Reason'] = 'UNKNOWN'
                
                # Hold Time (Minutes)
                if 'entry_time' in trades_df.columns and 'exit_time' in trades_df.columns:
                    entry_dt = pd.to_datetime(trades_df['entry_time'])
                    exit_dt = pd.to_datetime(trades_df['exit_time'])
                    formatted_trades['Hold Time (Minutes)'] = ((exit_dt - entry_dt).dt.total_seconds() / 60).round(2)
                
                # Write to Excel
                formatted_trades.to_excel(writer, sheet_name='Trades', index=False)
                format_excel_sheet(writer, 'Trades', formatted_trades)
                logger.info(f"Exported {len(formatted_trades)} trades to Excel")
            else:
                # Empty trades
                trades_df.to_excel(writer, sheet_name='Trades', index=False)
                format_excel_sheet(writer, 'Trades', trades_df)
        
        # 3. Equity Curve Sheet
        if equity_path.exists():
            equity_df = pd.read_csv(equity_path)
            
            # Convert timestamps to IST
            timestamp_cols = ['timestamp']
            equity_df = convert_to_ist(equity_df, timestamp_cols)
            
            # Format numeric columns
            numeric_cols = ['equity', 'cash', 'market_value', 'unrealized_pnl', 'realized_pnl']
            for col in numeric_cols:
                if col in equity_df.columns:
                    equity_df[col] = equity_df[col].round(2)
            
            equity_df.to_excel(writer, sheet_name='Equity Curve', index=False)
            format_excel_sheet(writer, 'Equity Curve', equity_df)
            logger.info(f"Exported {len(equity_df)} equity curve points to Excel")
        
        # 4. Positions Sheet (optional)
        if include_positions and positions_path.exists():
            positions_df = pd.read_csv(positions_path)
            
            # Convert timestamps to IST
            timestamp_cols = ['timestamp', 'entry_ts']
            positions_df = convert_to_ist(positions_df, timestamp_cols)
            
            # Format numeric columns
            numeric_cols = ['avg_price', 'mtm_price', 'unrealized_pnl', 'realized_pnl']
            for col in numeric_cols:
                if col in positions_df.columns:
                    positions_df[col] = positions_df[col].round(2)
            
            positions_df.to_excel(writer, sheet_name='Positions', index=False)
            format_excel_sheet(writer, 'Positions', positions_df)
            logger.info(f"Exported {len(positions_df)} position snapshots to Excel")
        
        # 5. Selection Sheet (optional, for debugging)
        if include_selection and selection_path.exists():
            selection_df = pd.read_csv(selection_path)
            
            # Convert timestamps to IST
            timestamp_cols = ['timestamp']
            selection_df = convert_to_ist(selection_df, timestamp_cols)
            
            # Format numeric columns
            numeric_cols = ['spot', 'strike', 'delta', 'bid', 'ask', 'mid', 'spread_pct']
            for col in numeric_cols:
                if col in selection_df.columns:
                    selection_df[col] = selection_df[col].round(4)
            
            selection_df.to_excel(writer, sheet_name='Contract Selection', index=False)
            format_excel_sheet(writer, 'Contract Selection', selection_df)
            logger.info(f"Exported {len(selection_df)} contract selections to Excel")
    
    logger.info(f"Excel export complete: {output_path}")
    return output_path


def export_run_to_excel(run_id: str, runs_root: Path = Path("runs")) -> Path:
    """
    Export a specific run to Excel by run ID.
    
    Args:
        run_id: Run ID (e.g., "run-20260111-104522-948")
        runs_root: Root directory containing run folders
        
    Returns:
        Path to created Excel file
    """
    run_dir = runs_root / run_id
    
    if not run_dir.exists():
        raise FileNotFoundError(f"Run directory not found: {run_dir}")
    
    return export_to_excel(run_dir)


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) > 1:
        run_id = sys.argv[1]
        output_path = export_run_to_excel(run_id)
        print(f"Exported to: {output_path}")
    else:
        print("Usage: python excel_export.py <run_id>")
        print("Example: python excel_export.py run-20260111-104522-948")
